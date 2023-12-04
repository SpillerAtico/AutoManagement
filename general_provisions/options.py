import openpyxl
from typing import List
from config import attachments, your_option
from view import text_collection
from utils.enums import AttachmentsTable8, AttachmentsTable6

workbook = openpyxl.load_workbook(attachments)
sheets = openpyxl.load_workbook(attachments).sheetnames


def get_table_options_3(option: int, sheet) -> str:
    options = {cell.coordinate: cell.value for content in sheet['A5':'A94'] for cell in content if
               cell.value is not None}

    key_list, val_list = list(options.keys()), list(options.values())
    return key_list[val_list.index(option)]


def get_ships_table_2(ships: list, sheet) -> dict:
    options = {str(cell.coordinate): str(cell.value) for content in sheet['B6':'B23'] for cell in content if
               cell.value is not None}

    key_list, val_list = list(options.keys()), list(options.values())

    return {str(sheet[key_list[val_list.index(ship)]].value): str(key_list[val_list.index(ship)]) for ship in ships}


def your_ships(option: int):
    workbook.active = workbook[sheets[2]]  # Приложение 3
    sheet = workbook.active
    cell_option = get_table_options_3(option, sheet)

    your_ship = get_ships(cell_option, sheet)
    your_ships = [your_ship[0], your_ship[1], your_ship[2]]
    return your_ships


def get_ships_table_8(ships: list, sheet) -> dict:
    options = {str(cell.coordinate): str(cell.value) for content in sheet['B4':'S4'] for cell in content if
               cell.value is not None}

    key_list, val_list = list(options.keys()), list(options.values())

    return {str(sheet[key_list[val_list.index(ship)]].value): str(key_list[val_list.index(ship)]) for ship in ships}


def get_ships(cell_option, sheet) -> List[str]:
    ship_cells = [cell_option.replace('A', order) for order in ('E', 'F', 'G')]
    ships = [str(sheet[ship].value) for ship in ship_cells]

    return ships


def get_cargos(cell_option, sheet) -> List[str]:
    cargo_forward = cell_option.replace('A', 'D')

    old_num_option = cargo_forward[1:]
    new_num_option = int(cargo_forward[1:]) + 1
    # word_cell_option = type_cargo_1[:1]

    cargo_reverse = cargo_forward.replace(old_num_option, str(new_num_option))

    cargo = [str(sheet[cargo_forward].value), str(sheet[cargo_reverse].value)]

    return cargo


def get_ports(option) -> list:
    workbook.active = workbook[sheets[2]]  # Приложение 3
    sheet = workbook.active
    cell_option = get_table_options_3(option, sheet)

    cell_ports = [cell_option.replace('A', order) for order in ('B', 'C')]

    ports = [str(sheet[port].value) for port in cell_ports]

    return ports


def get_type_ship(cell_option, sheet):
    cell_ports = [cell_option.replace('B', 'A')]
    type = [str(sheet[port].value) for port in cell_ports]

    return type


def get_distance(cell_option, sheet) -> str:
    cell_distance = cell_option.replace('A', 'H')

    distance = str(sheet[cell_distance].value)

    return distance


def get_attachments_8(ship_cells, sheet, enum):
    ships = list(ship_cells.values())
    option = {str(sheet[ship].value): str(sheet[ship.replace('4', enum)].value) for ship in ships}

    return option


def cargo_for_calc(option):
    workbook.active = workbook[sheets[2]]  # Приложение 3
    sheet = workbook.active
    cell_option = get_table_options_3(option, sheet)

    cargo_forward = cell_option.replace('A', 'D')

    old_num_option = cargo_forward[1:]
    new_num_option = int(cargo_forward[1:]) + 1

    cargo_reverse = cargo_forward.replace(old_num_option, str(new_num_option))

    cargo = {text_collection.load_volume_f: str(sheet[cargo_forward].value),
             text_collection.load_volume_r: str(sheet[cargo_reverse].value)}

    return cargo


def get_info_ships(option: int, enum) -> dict:
    workbook.active = workbook[sheets[8]]  # Приложение 8, теперь взаимодействуем с полученными корабликами
    sheet = workbook.active

    your_ship = your_ships(option)
    ship_cells = get_ships_table_8(your_ship, sheet)

    ships = list(ship_cells.values())
    data = {str(sheet[ship].value): str(sheet[ship.replace('4', enum)].value) for ship in ships}

    return data


def get_info_distance(option: int):
    workbook.active = workbook[sheets[2]]  # Приложение 3
    sheet = workbook.active
    cell_option = get_table_options_3(option, sheet)

    distance = get_distance(cell_option, sheet)
    return distance


def get_info_ships_6(option: int, word):
    workbook.active = workbook[sheets[6]]  # Приложение 6
    sheet = workbook.active
    options = {str(cell.coordinate): str(cell.value) for content in sheet['B4':'B21'] for cell in content if
               str(cell.value) in your_ships(option)}
    keys = list(options.keys())

    data = {str(sheet[cell_ship].value): str(sheet[cell_ship.replace('B', word)].value)
            for cell_ship in keys}

    return data


def get_ports_for_calc(option) -> dict:
    workbook.active = workbook[sheets[2]]  # Приложение 3
    sheet = workbook.active
    cell_option = get_table_options_3(option, sheet)

    ports_cells = [str(sheet[cell_option.replace('A', order)].value) for order in ('B', 'C')]

    port_1 = ports_cells[0][ports_cells[0].index('(') + 1: ports_cells[0].index(')')]
    port_2 = ports_cells[1][ports_cells[1].index('(') + 1: ports_cells[1].index(')')]

    ports = {ship: (port_1, port_2) for ship in your_ships(option)}

    return ports


def ship_fees_table_7(option: int):
    workbook.active = workbook[sheets[7]]  # Приложение 7
    sheet = workbook.active

    ships, ports = list(get_ports_for_calc(option).keys()), list(get_ports_for_calc(option).values())[0]

    ports_cells_7 = [str(cell.coordinate)[:1] for content in sheet['B3':'S3'] for cell in content
                     if str(cell.value) in ports]

    ships_cells_7 = {str(cell.value): str(cell.coordinate) for content in sheet['A5':'A22'] for cell in content
                     if str(cell.value) in ships}

    list_cell_ship = list(ships_cells_7.values())
    for i in range(3):
        ships_cells_7[ships[i]] = [float(sheet[list_cell_ship[i].replace('A', word)].value)
                                   for word in ports_cells_7]

    return ships_cells_7


def get_attachments_2(ship_cells, sheet, enum):
    ships = list(ship_cells.values())
    option = {str(sheet[ship].value): str(sheet[ship.replace('B', enum)].value) for ship in ships}

    return option
